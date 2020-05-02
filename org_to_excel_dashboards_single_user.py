
from arcgis import gis
from openpyxl import load_workbook

if __name__ == "__main__":

    GIS_USER = ""
    USERNAME = "USERNAME"
    WB = load_workbook('name.xlsx') # load existing Microsoft Excel document
    WS = WB.active

    ITEMS = GIS_USER.content.search(f"type:Dashboard owner:{USERNAME}", max_items=100000, outside_org=False) # find Dashboard items in your org
    print(f"I have {len(ITEMS)} items to work on")
    ALL_ITEMS = []
    for INDEX, ITEM in enumerate(ITEMS):
        print(f"working on Item {INDEX}")
        DASHBOARD_DATA = ITEM.get_data()
        if DASHBOARD_DATA != {}: # Sometimes dashboard's are empty. We don't want those ones!
            ITEMS_WITH_VERSION = (ITEM, DASHBOARD_DATA['version']) 
            ALL_ITEMS.append((ITEMS_WITH_VERSION))

    for ITEM_INDEX, row in enumerate(WS.iter_rows(min_row=2, max_row=len(ALL_ITEMS)+1)): # start at the second row and keep adding rows until you run out of items.
        row[0].value = ALL_ITEMS[ITEM_INDEX][0].title
        row[1].value = ALL_ITEMS[ITEM_INDEX][0].id
        row[2].value = ALL_ITEMS[ITEM_INDEX][0].type
        row[3].value = ALL_ITEMS[ITEM_INDEX][1]


    WB.save('name.xlsx')