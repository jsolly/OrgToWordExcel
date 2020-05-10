from openpyxl import load_workbook
from other.my_secrets import REGRESSION_DEVEXT_DBQA_GIS


def get_items_from_folder(
    gis_obj, folder, item_types=None
) -> list:  # folder=None returns the root folder
    items = gis_obj.users.me.items(folder=folder)

    if item_types:
        items = [item for item in items if item.type in item_types]

    return items


if __name__ == "__main__":
    WB = load_workbook("name.xlsx")
    WS = WB.active
    FOLDERS = (
        REGRESSION_DEVEXT_DBQA_GIS.users.me.folders
    )  # Get a list of all folders in AGOL
    print(f"I have {len(FOLDERS)} folders to work on")

    ALL_ITEMS = []
    for FOLDER_INDEX, FOLDER in enumerate(FOLDERS):
        print(f"I am working on Folder {FOLDER_INDEX} with name, {FOLDER['title']}")
        FOLDER_ITEMS = get_items_from_folder(
            gis_obj=REGRESSION_DEVEXT_DBQA_GIS, folder=FOLDER
        )
        ITEMS_WITH_FOLDER = [(ITEM, FOLDER["title"]) for ITEM in FOLDER_ITEMS]
        ALL_ITEMS += ITEMS_WITH_FOLDER

    for ITEM_INDEX, row in enumerate(
        WS.iter_rows(min_row=2, max_row=len(ALL_ITEMS) + 1)
    ):  # start at the second row and keep adding rows until you run out of items
        row[0].value = ALL_ITEMS[ITEM_INDEX][0].title
        row[1].value = ALL_ITEMS[ITEM_INDEX][0].id
        row[2].value = ALL_ITEMS[ITEM_INDEX][1]
        row[3].value = ALL_ITEMS[ITEM_INDEX][0].type
        row[
            4
        ].value = f"""=HYPERLINK("{REGRESSION_DEVEXT_DBQA_GIS.url}//home/item.html?id={ALL_ITEMS[ITEM_INDEX][0].id}", "Open A copy in Develop 4")"""  # Change to item details page
        row[4].style = "Hyperlink"

    WB.save("name.xlsx")
